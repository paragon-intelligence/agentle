"""
Excel Spreadsheet Parser Module

This module provides functionality for parsing Excel spreadsheet files (.xls, .xlsx) into
structured document representations. It extracts text content, tables, and images from
spreadsheets while preserving their structure.
"""

from collections.abc import MutableSequence
from pathlib import Path
from typing import Literal, override


from rsb.functions.ext2mime import ext2mime


from agentle.agents.agent import Agent
from agentle.generations.models.message_parts.file import FilePart
from agentle.generations.models.structured_outputs_store.visual_media_description import (
    VisualMediaDescription,
)
from agentle.parsing.document_parser import DocumentParser
from agentle.parsing.factories.visual_description_agent_default_factory import (
    visual_description_agent_default_factory,
)
from agentle.parsing.image import Image
from agentle.parsing.page_item.table_page_item import TablePageItem
from agentle.parsing.parsed_document import ParsedDocument
from agentle.parsing.section_content import SectionContent
from rsb.models.field import Field


class XlsxFileParser(DocumentParser):
    """
    Parser for processing Excel spreadsheet files (.xls, .xlsx).

    This parser extracts content from Excel files, including text data, tables, and
    embedded images. Each worksheet in the Excel file is parsed as a separate section
    in the resulting ParsedDocument. The parser preserves the tabular structure of
    the data by creating TablePageItem objects for each worksheet.

    **Attributes:**

    *   `strategy` (Literal["high", "low"]):
        The parsing strategy to use. Defaults to "high".
        - "high": Performs thorough parsing including image analysis
        - "low": Performs basic parsing without analyzing embedded images

        **Example:**
        ```python
        parser = XlsxFileParser(strategy="low")  # Use faster, less intensive parsing
        ```

    *   `visual_description_agent` (Agent[VisualMediaDescription]):
        An optional custom agent for visual media description. If provided and strategy
        is "high", this agent will be used to analyze images embedded in the spreadsheet.
        Defaults to the agent created by `visual_description_agent_default_factory()`.

        **Example:**
        ```python
        from agentle.agents.agent import Agent
        from agentle.generations.models.structured_outputs_store.visual_media_description import VisualMediaDescription

        custom_agent = Agent(
            model="gemini-2.0-pro-vision",
            instructions="Focus on chart and diagram analysis in spreadsheets",
            response_schema=VisualMediaDescription
        )

        parser = XlsxFileParser(visual_description_agent=custom_agent)
        ```

    **Usage Examples:**

    Basic parsing of an Excel file:
    ```python
    from agentle.parsing.parsers.xlsx import XlsxFileParser

    # Create a parser with default settings
    parser = XlsxFileParser()

    # Parse an Excel file
    parsed_doc = parser.parse("financial_data.xlsx")

    # Access the worksheets (as sections)
    for section in parsed_doc.sections:
        print(f"Worksheet: {section.number}")

        # Access table data
        for item in section.items:
            if isinstance(item, TablePageItem):
                print(f"Table with {len(item.rows)} rows")
                # Get CSV representation
                print(item.csv)
    ```

    Parsing with a custom visual description agent:
    ```python
    from agentle.agents.agent import Agent
    from agentle.generations.models.structured_outputs_store.visual_media_description import VisualMediaDescription
    from agentle.generations.providers.google.google_genai_generation_provider import GoogleGenaiGenerationProvider

    # Create a custom agent for chart analysis
    chart_analysis_agent = Agent(
        model="gemini-2.0-pro-vision",
        instructions="Analyze charts and graphs in spreadsheets with focus on trends and insights",
        generation_provider=GoogleGenaiGenerationProvider(),
        response_schema=VisualMediaDescription
    )

    # Create a parser with the custom agent
    parser = XlsxFileParser(
        strategy="high",
        visual_description_agent=chart_analysis_agent
    )

    # Parse a spreadsheet with charts
    report = parser.parse("quarterly_report.xlsx")

    # Access chart descriptions
    for section in report.sections:
        for image in section.images:
            if image.ocr_text:
                print(f"Chart text: {image.ocr_text}")
    ```
    """

    type: Literal["xlsx"] = "xlsx"

    strategy: Literal["high", "low"] = Field(default="high")

    visual_description_agent: Agent[VisualMediaDescription] = Field(
        default_factory=visual_description_agent_default_factory,
    )
    """
    The agent to use for generating the visual description of the document.
    Useful when you want to customize the prompt for the visual description.
    """

    @override
    async def parse_async(
        self,
        document_path: str,
    ) -> ParsedDocument:
        """
        Asynchronously parse an Excel spreadsheet file and convert it to a structured representation.

        This method reads an Excel file, extracts each worksheet as a separate section,
        and processes any embedded images found in the worksheets. For each worksheet,
        it creates a TablePageItem containing the structured table data and a CSV representation.

        Args:
            document_path (str): Path to the Excel file to be parsed

        Returns:
            ParsedDocument: A structured representation of the Excel file where:
                - Each worksheet is a separate section
                - Each section contains a TablePageItem with the worksheet data
                - Images are extracted and (optionally) analyzed

        Example:
            ```python
            import asyncio
            from agentle.parsing.parsers.xlsx import XlsxFileParser

            async def process_excel():
                parser = XlsxFileParser(strategy="high")
                result = await parser.parse_async("data.xlsx")

                # Print worksheet names and row counts
                for i, section in enumerate(result.sections):
                    print(f"Worksheet {i+1}")

                    # Find table items
                    for item in section.items:
                        if hasattr(item, 'rows'):
                            print(f"  - Contains {len(item.rows)} rows of data")

            asyncio.run(process_excel())
            ```
        """
        import csv
        import io

        from openpyxl import Workbook, load_workbook

        wb: Workbook = load_workbook(document_path, data_only=True)
        sections: MutableSequence[SectionContent] = []

        for sheet_index, sheet in enumerate(wb.worksheets, start=1):
            # Gather structured data
            rows: list[list[str]] = []
            row_texts: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                # Process cell values
                cell_values = [str(cell) if cell is not None else "" for cell in row]
                rows.append(cell_values)
                row_texts.append("\t".join(cell_values))

            combined_text = "\n".join(row_texts)

            # Generate CSV content
            csv_buffer = io.StringIO()
            csv_writer = csv.writer(csv_buffer)
            csv_writer.writerows(rows)
            csv_str = csv_buffer.getvalue().strip()

            # Process images
            sheet_images: list[tuple[str, bytes]] = []
            if hasattr(sheet, "_images"):
                image_list = getattr(sheet, "_images", [])
                for img_idx, img in enumerate(image_list, start=1):
                    img_data = getattr(img, "_data", None)
                    if img_data is not None:
                        image_name = f"{sheet.title}_img_{img_idx}.png"
                        sheet_images.append((image_name, img_data))

            final_images: MutableSequence[Image] = []
            # Generate image descriptions if needed
            if self.visual_description_agent and self.strategy == "high":
                image_descriptions: MutableSequence[str] = []
                for img_idx, image_obj in enumerate(sheet_images, start=1):
                    agent_input = FilePart(
                        mime_type=ext2mime(Path(image_obj[0]).suffix),
                        data=image_obj[1],
                    )
                    agent_response = await self.visual_description_agent.run_async(
                        agent_input
                    )
                    image_md = agent_response.parsed.md
                    image_descriptions.append(
                        f"Worksheet {sheet.title} - Image {img_idx}: {image_md}"
                    )
                    final_images.append(
                        Image(
                            name=image_obj[0],
                            contents=image_obj[1],
                            ocr_text=agent_response.parsed.ocr_text,
                        )
                    )

                if image_descriptions:
                    combined_text += "\n\n" + "\n".join(image_descriptions)

            # Create table page item
            table_item = TablePageItem(rows=rows, csv=csv_str, is_perfect_table=True)

            section_content = SectionContent(
                number=sheet_index,
                text=combined_text,
                md=combined_text,
                images=final_images,
                items=[table_item],
            )
            sections.append(section_content)

        return ParsedDocument(
            name=Path(document_path).name,
            sections=sections,
        )
